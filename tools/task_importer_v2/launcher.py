import http.server
import base64
import json
import os
import re
import socket
import socketserver
import sys
import threading
import time
import unicodedata
import webbrowser
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl


APP_NAME = "Datacora JM 2.0"
HOST = "127.0.0.1"
PORT = 8092
TASK_TYPES = ["Plan Preventivo Mantencion", "DT", "Mutualidad", "Emergencia", "Acta", "Seremi", "SEC"]
COLUMN_ALIASES = {
    "rbd": ["rbd", "codigo rbd", "cod rbd", "establecimiento rbd"],
    "technician": ["nombre tecnico", "nombre tecnico", "tecnico", "correo tecnico", "email tecnico", "tecnico email", "assigned to", "asignado a"],
    "due_date": ["fecha planificada", "fecha", "fecha tarea", "fecha vencimiento", "fecha visita", "due date"],
    "task_type": ["tipo visita", "tipo", "tipo tarea", "tipo de tarea", "tipo formulario", "formulario", "visita", "task type"],
    "description": ["descripcion", "observacion", "detalle", "comentario"],
    "priority": ["prioridad", "priority"],
    "status": ["estado", "status"],
}


def app_dir():
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS")) / "task_importer_v2"
    return Path(__file__).resolve().parent


def port_is_open(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
        return probe.connect_ex((HOST, port)) == 0


def normalize(value):
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def map_headers(headers):
    by_normalized = {normalize(header): header for header in headers.values()}
    mapped = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            header = by_normalized.get(normalize(alias))
            if header:
                mapped[target] = header
                break
    return mapped


def find_user(users, value):
    key = normalize(value)
    for user in users:
        if normalize(user.get("email")) == key:
            return user
    matches = [user for user in users if normalize(user.get("full_name")) == key]
    return matches[0] if len(matches) == 1 else None


def find_establishment(establishments, rbd):
    key = clean_text(rbd)
    for establishment in establishments:
        if clean_text(establishment.get("rbd")) == key:
            return establishment
    return None


def parse_workbook(content_bytes, users, establishments):
    workbook = openpyxl.load_workbook(BytesIO(content_bytes), data_only=True)
    sheet = workbook.active
    header_values = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = {idx: header for idx, header in enumerate(header_values) if header}
    mapped = map_headers(headers)
    missing = [name for name in ("rbd", "technician") if name not in mapped]
    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing))

    rows = []
    today = date.today()
    for excel_row in sheet.iter_rows(min_row=2):
        raw_by_header = {headers[idx]: excel_row[idx].value for idx in headers if idx < len(excel_row)}
        if not any(clean_text(value) for value in raw_by_header.values()):
            continue

        raw = {key: raw_by_header.get(header) for key, header in mapped.items()}
        errors = []
        rbd = clean_text(raw.get("rbd"))
        technician_value = clean_text(raw.get("technician"))
        due_date = as_iso_date(raw.get("due_date")) or today.isoformat()
        task_type = clean_text(raw.get("task_type")) or "Plan Preventivo Mantencion"
        priority = normalize(raw.get("priority")) or "media"
        status = normalize(raw.get("status")) or "pendiente"

        if not rbd:
            errors.append("RBD requerido.")
        if not technician_value:
            errors.append("Tecnico requerido.")
        if priority not in ("alta", "media", "baja"):
            errors.append("Prioridad invalida.")
        if status not in ("pendiente", "urgente", "completada", "cancelada"):
            errors.append("Estado invalido.")
        try:
            parsed_due_date = date.fromisoformat(due_date)
            if parsed_due_date < today:
                errors.append("Fecha anterior a la actual.")
        except ValueError:
            errors.append("Fecha invalida.")

        establishment = find_establishment(establishments, rbd) if rbd else None
        if rbd and not establishment:
            errors.append(f"RBD {rbd} no corresponde a la zona del JM.")

        technician = find_user(users, technician_value)
        if technician_value and not technician:
            errors.append(f"Tecnico no encontrado: {technician_value}")

        payload = {
            "rbd": rbd,
            "establishmentId": establishment.get("id") if establishment else "",
            "assignedTo": technician.get("id") if technician else "",
            "dueDate": due_date,
            "type": task_type,
            "description": clean_text(raw.get("description")),
            "priority": priority,
            "status": status,
        }
        rows.append({
            "number": excel_row[0].row,
            "status": "Valida" if not errors else "Error",
            "errors": errors,
            "payload": payload,
            "rbd": rbd,
            "establishment": establishment.get("name") if establishment else "",
            "technician": technician.get("full_name") if technician else technician_value,
            "dueDate": due_date,
            "type": task_type,
            "priority": priority,
            "description": payload["description"],
        })
    return rows


class QuietHandler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        super().end_headers()

    def log_message(self, _format, *_args):
        return

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def do_POST(self):
        if self.path != "/local/validate-xlsx":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            payload = json.loads(body.decode("utf-8"))
            file_bytes = base64.b64decode(payload.get("fileBase64") or "")
            rows = parse_workbook(file_bytes, payload.get("users") or [], payload.get("establishments") or [])
            response = {
                "rows": rows,
                "total": len(rows),
                "valid": sum(1 for row in rows if not row["errors"]),
                "invalid": sum(1 for row in rows if row["errors"]),
            }
            self.send_json(200, response)
        except Exception as exc:
            self.send_json(400, {"error": str(exc)})

    def send_json(self, status, payload):
        content = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)


def edge_path():
    candidates = [
        os.environ.get("ProgramFiles", "") + r"\Microsoft\Edge\Application\msedge.exe",
        os.environ.get("ProgramFiles(x86)", "") + r"\Microsoft\Edge\Application\msedge.exe",
        os.environ.get("LocalAppData", "") + r"\Microsoft\Edge\Application\msedge.exe",
    ]
    return next((path for path in candidates if path and Path(path).exists()), "")


def main():
    root = app_dir()
    if not (root / "index.html").exists():
        raise RuntimeError(f"No se encontró la interfaz de {APP_NAME}.")

    os.chdir(root)
    url = f"http://{HOST}:{PORT}/"

    if port_is_open(PORT):
        open_app_window(url)
        return

    server = socketserver.TCPServer((HOST, PORT), QuietHandler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()

    open_app_window(url)

    while True:
        time.sleep(3600)


def open_app_window(url):
    edge = edge_path()
    if edge:
        os.startfile(edge, arguments=f"--app={url} --new-window")
    else:
        webbrowser.open(url)


if __name__ == "__main__":
    main()

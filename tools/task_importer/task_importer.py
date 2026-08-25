import json
import os
import queue
import re
import threading
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, TOP, X, BooleanVar, StringVar, Tk, filedialog, messagebox
from tkinter import ttk

import openpyxl
import requests


APP_TITLE = "Datacora - Importador masivo de tareas"
CONFIG_DIR = Path(os.environ.get("APPDATA", str(Path.home()))) / "Datacora"
CONFIG_FILE = CONFIG_DIR / "task_importer_config.json"
DEFAULT_API_URL = os.environ.get("DATACORA_API_URL") or "http://" + "127.0.0.1" + ":8081"

COLOR_BG = "#f4f7f6"
COLOR_SURFACE = "#ffffff"
COLOR_PRIMARY = "#0b6a43"
COLOR_PRIMARY_DARK = "#075536"
COLOR_TEXT = "#102033"
COLOR_MUTED = "#667381"
COLOR_BORDER = "#d9e3df"
COLOR_SOFT = "#e7f4ee"

COLUMN_ALIASES = {
    "rbd": ["rbd", "codigo rbd", "cod rbd", "establecimiento rbd"],
    "technician": ["nombre tecnico", "nombre técnico", "nombre t?cnico", "tecnico", "técnico", "t?cnico", "correo tecnico", "correo técnico", "correo t?cnico", "email tecnico", "email técnico", "email t?cnico", "tecnico email", "técnico email", "t?cnico email", "assigned to", "asignado a"],
    "due_date": ["fecha planificada", "fecha", "fecha tarea", "fecha vencimiento", "fecha visita", "due date"],
    "task_type": ["tipo visita", "tipo", "tipo tarea", "tipo de tarea", "tipo formulario", "formulario", "visita", "task type"],
    "description": ["descripcion", "descripción", "observacion", "observación", "detalle", "comentario"],
    "priority": ["prioridad", "priority"],
    "status": ["estado", "status"],
    "required_sections": ["secciones requeridas", "required sections"],
    "critical_sections": ["secciones criticas", "secciones críticas", "critical sections"],
    "section_minimums": ["minimos secciones", "mínimos secciones", "section minimums"]
}

MOBILE_TEMPLATE_HEADERS = ["Fecha planificada", "Nombre TÃ©cnico", "RBD", "Tipo Visita", "Prioridad"]
MOBILE_TASK_TYPES = ["Plan Preventivo MantenciÃ³n", "DT", "Mutualidad", "Emergencia", "Acta", "Seremi", "SEC"]
MOBILE_PRIORITIES = ["media", "alta", "baja"]

SECTION_ALIASES = {
    "calor": "heat",
    "heat": "heat",
    "electricidad": "electricity",
    "electricity": "electricity",
    "frio": "cold",
    "frío": "cold",
    "cold": "cold",
    "vectores": "vectors",
    "vectors": "vectors",
    "agua": "water",
    "water": "water",
    "infraestructura": "infrastructure",
    "infrastructure": "infrastructure",
    "encargado pae": "pae-manager",
    "pae": "pae-manager",
    "mpa": "mpa",
    "patio servicio": "service-yard",
    "service yard": "service-yard",
    "verificadores rbd": "rbd-checkers",
    "rbd checkers": "rbd-checkers",
}


def normalize(value):
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_iso_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def split_list(value):
    return [item.strip() for item in re.split(r"[;,|]", clean_text(value)) if item.strip()]


def normalize_sections(value):
    sections = []
    for item in split_list(value):
        key = normalize(item)
        sections.append(SECTION_ALIASES.get(key, key))
    return sections


def parse_minimums(value):
    minimums = {}
    for item in split_list(value):
      match = re.match(r"^\s*([^:=]+)\s*[:=]\s*(\d+)\s*$", item)
      if match:
          section = SECTION_ALIASES.get(normalize(match.group(1)), normalize(match.group(1)))
          minimums[section] = int(match.group(2))
    return minimums


@dataclass
class ImportRow:
    number: int
    raw: dict
    payload: dict = field(default_factory=dict)
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    status: str = "Pendiente"
    result: str = ""


class DatacoraClient:
    def __init__(self, base_url):
        self.base_url = base_url.rstrip("/")
        self.token = ""
        self.user = None

    def headers(self):
        return {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}

    def login(self, email, password):
        try:
            response = requests.post(
                f"{self.base_url}/auth/login",
                json={"email": email, "password": password},
                timeout=20
            )
        except requests.RequestException:
            raise RuntimeError("No se pudo conectar con Datácora. Verifica que el backend esté iniciado.")
        data = self._json(response)
        self.token = data["access_token"]
        self.user = data.get("user") or {}
        if not (self.user.get("canAssignTasks") or self.user.get("canManageUsers")):
            raise RuntimeError("El usuario no tiene permisos para crear tareas.")
        return self.user

    def get_users(self):
        response = requests.get(f"{self.base_url}/api/users", headers=self.headers(), timeout=30)
        return self._json(response)

    def get_tasks(self):
        response = requests.get(f"{self.base_url}/api/tasks", headers=self.headers(), timeout=30)
        return self._json(response)

    def get_establishments(self, search="", branch_id=""):
        response = requests.get(
            f"{self.base_url}/api/establishments",
            headers=self.headers(),
            params={"search": search, "branchId": branch_id or ""},
            timeout=30
        )
        return self._json(response)

    def search_establishment(self, rbd):
        response = requests.get(
            f"{self.base_url}/api/establishments",
            headers=self.headers(),
            params={"search": str(rbd)},
            timeout=30
        )
        rows = self._json(response)
        exact = [row for row in rows if str(row.get("rbd", "")).strip() == str(rbd).strip()]
        return exact[0] if exact else None

    def create_task(self, payload):
        response = requests.post(f"{self.base_url}/api/tasks", headers=self.headers(), json=payload, timeout=30)
        return self._json(response)

    @staticmethod
    def _json(response):
        try:
            data = response.json()
        except Exception:
            data = {"error": response.text}
        if not response.ok:
            raise RuntimeError(data.get("error") or data.get("message") or response.text or f"HTTP {response.status_code}")
        return data


class TaskImporterApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1440x820")
        self.root.minsize(1180, 720)
        self.root.configure(bg=COLOR_BG)
        self.queue = queue.Queue()
        self.client = None
        self.users = []
        self.tasks = []
        self.establishments = []
        self.rows = []
        self.excel_path = StringVar()
        self.api_url = StringVar(value=DEFAULT_API_URL)
        self.email = StringVar()
        self.password = StringVar()
        self.only_valid = BooleanVar(value=True)
        self.status = StringVar(value="Inicia sesión para continuar.")
        self.session_label = StringVar(value="")
        self.valid_count = StringVar(value="0")
        self.error_count = StringVar(value="0")
        self.created_count = StringVar(value="0")
        self.pending_count = StringVar(value="0")
        self.preview_title = StringVar(value="Sin planilla cargada")
        self.preview_detail = StringVar(value="Descarga la plantilla, completa las filas y valida antes de crear tareas.")
        self.page_title = StringVar(value="Crear tarea masiva")
        self.page_subtitle = StringVar(value="Carga una planilla Excel y asigna visitas a técnicos de forma rápida.")
        self.current_pending_count = StringVar(value="0")
        self.current_active_count = StringVar(value="0")
        self.current_urgent_count = StringVar(value="0")
        self.current_completed_count = StringVar(value="0")
        self.current_technicians_count = StringVar(value="0")
        self.current_total_count = StringVar(value="0")
        self._load_config()
        self._setup_styles()
        self._build_ui()
        self._poll_queue()

    def _load_config(self):
        if CONFIG_FILE.exists():
            try:
                data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
                self.email.set(data.get("email") or "")
            except Exception:
                pass

    def _save_config(self):
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_FILE.write_text(json.dumps({
            "email": self.email.get().strip()
        }, indent=2), encoding="utf-8")

    def _build_ui(self):
        top = ttk.Frame(self.root, padding=12)
        top.pack(fill=X)

        ttk.Label(top, text="Correo").pack(side=LEFT)
        ttk.Entry(top, textvariable=self.email, width=34).pack(side=LEFT, padx=(6, 14))
        ttk.Label(top, text="Clave").pack(side=LEFT)
        ttk.Entry(top, textvariable=self.password, width=22, show="*").pack(side=LEFT, padx=(6, 14))
        ttk.Button(top, text="Conectar", command=self.login).pack(side=LEFT)

        file_bar = ttk.Frame(self.root, padding=(12, 0, 12, 8))
        file_bar.pack(fill=X)
        ttk.Entry(file_bar, textvariable=self.excel_path).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(file_bar, text="Seleccionar Excel", command=self.choose_file).pack(side=LEFT, padx=8)
        ttk.Button(file_bar, text="Leer y validar", command=self.load_excel).pack(side=LEFT)
        ttk.Button(file_bar, text="Descargar plantilla", command=self.create_template).pack(side=LEFT, padx=(8, 0))

        actions = ttk.Frame(self.root, padding=(12, 0, 12, 8))
        actions.pack(fill=X)
        ttk.Checkbutton(actions, text="Crear solo filas válidas", variable=self.only_valid).pack(side=LEFT)
        ttk.Button(actions, text="Crear tareas", command=self.create_tasks).pack(side=LEFT, padx=8)
        ttk.Button(actions, text="Exportar resultado", command=self.export_results).pack(side=LEFT)
        ttk.Label(actions, textvariable=self.status).pack(side=RIGHT)

        columns = ("fila", "estado", "rbd", "establecimiento", "tecnico", "fecha", "tipo", "prioridad", "mensaje")
        self.tree = ttk.Treeview(self.root, columns=columns, show="headings", height=22)
        widths = {"fila": 60, "estado": 110, "rbd": 90, "establecimiento": 230, "tecnico": 230, "fecha": 100, "tipo": 190, "prioridad": 90, "mensaje": 310}
        for column in columns:
            self.tree.heading(column, text=column.capitalize())
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.pack(fill=BOTH, expand=True, padx=12, pady=(0, 8))

        log_frame = ttk.LabelFrame(self.root, text="Detalle")
        log_frame.pack(fill=BOTH, expand=False, padx=12, pady=(0, 12))
        self.log = ttk.Treeview(log_frame, columns=("mensaje",), show="headings", height=5)
        self.log.heading("mensaje", text="Mensaje")
        self.log.column("mensaje", width=1100)
        self.log.pack(fill=BOTH, expand=True)

    def choose_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if path:
            self.excel_path.set(path)

    def create_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Plantilla carga masiva tareas Datacora.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tareas"
        headers = ["RBD", "Correo técnico", "Fecha", "Tipo tarea", "Descripción", "Prioridad", "Estado", "Secciones requeridas", "Secciones críticas", "Mínimos secciones"]
        sheet.append(headers)
        sheet.append(["1", "tecnico.rm@soser.cl", date.today().isoformat(), "Plan Preventivo Mantención", "Visita programada", "media", "pendiente", "heat;electricity", "heat;electricity", "heat=2;electricity=6"])
        for index, width in enumerate([12, 32, 14, 28, 42, 14, 14, 28, 28, 28], start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        workbook.save(path)
        messagebox.showinfo(APP_TITLE, f"Plantilla creada:\n{path}")

    def login(self):
        self._run_background(self._login_worker)

    def _login_worker(self):
        self._set_status("Conectando...")
        client = DatacoraClient(self.api_url.get().strip())
        user = client.login(self.email.get().strip(), self.password.get())
        users = client.get_users()
        self.client = client
        self.users = users
        self._save_config()
        self._log(f"Conectado como {user.get('fullName') or user.get('email')}. Usuarios cargados: {len(users)}")
        self._set_status("Conectado.")

    def load_excel(self):
        if not self.client:
            messagebox.showwarning(APP_TITLE, "Primero conecta con usuario autorizado.")
            return
        self._run_background(self._load_excel_worker)

    def _load_excel_worker(self):
        path = self.excel_path.get().strip()
        if not path:
            raise RuntimeError("Selecciona una planilla Excel.")
        self._set_status("Leyendo Excel...")
        rows = self._parse_workbook(path)
        self._set_status("Validando contra SQL...")
        for row in rows:
            self._validate_row(row)
        self.rows = rows
        self._queue_ui(self._refresh_table)
        valid = sum(1 for row in rows if not row.errors)
        self._log(f"Validación lista. Filas: {len(rows)}. Válidas: {valid}. Con errores: {len(rows) - valid}.")
        self._set_status("Validación terminada.")

    def _parse_workbook(self, path):
        workbook = openpyxl.load_workbook(path, data_only=True)
        sheet = workbook.active
        header_values = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = {idx: header for idx, header in enumerate(header_values) if header}
        mapped = self._map_headers(headers)
        missing = [name for name in ("rbd", "technician") if name not in mapped]
        if missing:
            readable = ", ".join(missing)
            raise RuntimeError(f"Faltan columnas obligatorias: {readable}.")

        rows = []
        for excel_row in sheet.iter_rows(min_row=2):
            raw_by_header = {headers[idx]: excel_row[idx].value for idx in headers if idx < len(excel_row)}
            if not any(clean_text(value) for value in raw_by_header.values()):
                continue
            normalized_raw = {key: raw_by_header.get(header) for key, header in mapped.items()}
            rows.append(ImportRow(number=excel_row[0].row, raw=normalized_raw))
        return rows

    def _map_headers(self, headers):
        by_normalized = {normalize(header): header for header in headers.values()}
        mapped = {}
        for target, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                header = by_normalized.get(normalize(alias))
                if header:
                    mapped[target] = header
                    break
        return mapped

    def _validate_row(self, row):
        rbd = clean_text(row.raw.get("rbd"))
        technician_value = clean_text(row.raw.get("technician"))
        due_date = as_iso_date(row.raw.get("due_date")) or date.today().isoformat()
        task_type = clean_text(row.raw.get("task_type")) or "Plan Preventivo Mantención"
        priority = normalize(row.raw.get("priority")) or "media"
        status = normalize(row.raw.get("status")) or "pendiente"

        if not rbd:
            row.errors.append("RBD requerido.")
        if not technician_value:
            row.errors.append("Técnico requerido.")
        if priority not in ("alta", "media", "baja"):
            row.errors.append("Prioridad inválida.")
        if status not in ("pendiente", "urgente", "completada", "cancelada"):
            row.errors.append("Estado inválido.")

        establishment = None
        if rbd:
            try:
                establishment = self.client.search_establishment(rbd)
            except Exception as exc:
                row.errors.append(f"No se pudo validar RBD: {exc}")
            if not establishment:
                row.errors.append(f"RBD {rbd} no existe.")

        technician = self._find_user(technician_value)
        if not technician:
            row.errors.append(f"Técnico no encontrado: {technician_value}")

        section_config = {
            "requiredSections": normalize_sections(row.raw.get("required_sections")),
            "criticalSections": normalize_sections(row.raw.get("critical_sections")),
            "sectionMinimums": parse_minimums(row.raw.get("section_minimums")),
        }
        if not any(section_config.values()):
            section_config = {}

        row.payload = {
            "rbd": rbd,
            "assignedTo": technician.get("id") if technician else "",
            "dueDate": due_date,
            "type": task_type,
            "description": clean_text(row.raw.get("description")),
            "priority": priority,
            "status": status,
            "sectionConfig": section_config
        }
        row.raw["establishment_name"] = establishment.get("name") if establishment else ""
        row.raw["technician_name"] = technician.get("full_name") if technician else technician_value
        row.status = "Válida" if not row.errors else "Error"

    def _find_user(self, value):
        key = normalize(value)
        email_key = key
        for user in self.users:
            if normalize(user.get("email")) == email_key:
                return user
        matches = [user for user in self.users if normalize(user.get("full_name")) == key]
        if len(matches) == 1:
            return matches[0]
        return None

    def create_tasks(self):
        if not self.rows:
            messagebox.showwarning(APP_TITLE, "Primero lee y valida una planilla.")
            return
        self._run_background(self._create_tasks_worker)

    def _create_tasks_worker(self):
        rows = [row for row in self.rows if not row.errors] if self.only_valid.get() else self.rows
        created = 0
        failed = 0
        for row in rows:
            if row.errors:
                row.status = "Omitida"
                row.result = "Fila con errores de validación."
                failed += 1
                continue
            try:
                result = self.client.create_task(row.payload)
                row.status = "Creada"
                row.result = result.get("id", "Tarea creada")
                self.tasks.append({
                    "id": row.result,
                    "status": row.payload.get("status"),
                    "priority": row.payload.get("priority"),
                    "task_type": row.payload.get("type"),
                    "rbd": row.payload.get("rbd")
                })
                self._refresh_current_metrics()
                created += 1
            except Exception as exc:
                row.status = "Error"
                row.result = str(exc)
                failed += 1
            self._queue_ui(self._refresh_table)
        self._log(f"Carga terminada. Creadas: {created}. Errores/omitidas: {failed}.")
        self._set_status("Carga terminada.")

    def export_results(self):
        if not self.rows:
            messagebox.showwarning(APP_TITLE, "No hay resultados para exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="Resultado carga masiva tareas Datacora.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resultado"
        sheet.append(["Fila", "Estado", "RBD", "Establecimiento", "Técnico", "Fecha", "Tipo", "Prioridad", "Errores", "Resultado"])
        for row in self.rows:
            sheet.append([
                row.number,
                row.status,
                row.payload.get("rbd", ""),
                row.raw.get("establishment_name", ""),
                row.raw.get("technician_name", ""),
                row.payload.get("dueDate", ""),
                row.payload.get("type", ""),
                row.payload.get("priority", ""),
                "; ".join(row.errors),
                row.result,
            ])
        for index, width in enumerate([8, 14, 12, 34, 32, 14, 28, 14, 50, 45], start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        workbook.save(path)
        messagebox.showinfo(APP_TITLE, f"Resultado exportado:\n{path}")

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.rows:
            message = "; ".join(row.errors) if row.errors else row.result or "OK"
            self.tree.insert("", END, values=(
                row.number,
                row.status,
                row.payload.get("rbd", ""),
                row.raw.get("establishment_name", ""),
                row.raw.get("technician_name", ""),
                row.payload.get("dueDate", ""),
                row.payload.get("type", ""),
                row.payload.get("priority", ""),
                message
            ))

    def _run_background(self, fn):
        def runner():
            try:
                fn()
            except Exception as exc:
                message = str(exc) or "Ocurrió un error."
                self._log(message)
                self._set_status(message)
                self._queue_ui(lambda: messagebox.showerror(APP_TITLE, str(exc)))
        threading.Thread(target=runner, daemon=True).start()

    def _queue_ui(self, callback):
        self.queue.put(callback)

    def _poll_queue(self):
        while True:
            try:
                callback = self.queue.get_nowait()
            except queue.Empty:
                break
            callback()
        self.root.after(100, self._poll_queue)

    def _log(self, message):
        def append():
            self.log.insert("", END, values=(message,))
            self.log.yview_moveto(1)
        self._queue_ui(append)

    def _set_status(self, message):
        self._queue_ui(lambda: self.status.set(message))

    def _setup_styles(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("App.TFrame", background=COLOR_BG)
        style.configure("Surface.TFrame", background=COLOR_SURFACE)
        style.configure("Header.TFrame", background=COLOR_PRIMARY)
        style.configure("TLabel", background=COLOR_BG, foreground=COLOR_TEXT, font=("Segoe UI", 10))
        style.configure("Surface.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("Segoe UI", 10))
        style.configure("Muted.TLabel", background=COLOR_SURFACE, foreground=COLOR_MUTED, font=("Segoe UI", 9))
        style.configure("Brand.TLabel", background=COLOR_SURFACE, foreground=COLOR_PRIMARY, font=("Segoe UI", 24, "bold"))
        style.configure("Title.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("Segoe UI", 15, "bold"))
        style.configure("HeaderTitle.TLabel", background=COLOR_PRIMARY, foreground="#ffffff", font=("Segoe UI", 18, "bold"))
        style.configure("HeaderMuted.TLabel", background=COLOR_PRIMARY, foreground="#d8efe5", font=("Segoe UI", 9))
        style.configure("Primary.TButton", background=COLOR_PRIMARY, foreground="#ffffff", font=("Segoe UI", 10, "bold"), padding=(14, 8), borderwidth=0)
        style.map("Primary.TButton", background=[("active", COLOR_PRIMARY_DARK), ("pressed", COLOR_PRIMARY_DARK)])
        style.configure("Secondary.TButton", background=COLOR_SOFT, foreground=COLOR_PRIMARY, font=("Segoe UI", 10, "bold"), padding=(12, 7), borderwidth=0)
        style.map("Secondary.TButton", background=[("active", "#d9eee3"), ("pressed", "#d9eee3")])
        style.configure("TButton", font=("Segoe UI", 10), padding=(10, 6))
        style.configure("TEntry", padding=(8, 7), fieldbackground="#ffffff")
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=30, background="#ffffff", fieldbackground="#ffffff", foreground=COLOR_TEXT)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), foreground=COLOR_TEXT)
        style.configure("TLabelframe", background=COLOR_BG)
        style.configure("TLabelframe.Label", background=COLOR_BG, foreground=COLOR_TEXT, font=("Segoe UI", 10, "bold"))

    def _build_ui(self):
        self.login_frame = ttk.Frame(self.root, style="App.TFrame")
        self.workspace_frame = ttk.Frame(self.root, style="App.TFrame")
        self._build_login_view()
        self._build_workspace_view()
        self._show_login()

    def _build_login_view(self):
        outer = ttk.Frame(self.login_frame, style="App.TFrame", padding=32)
        outer.pack(fill=BOTH, expand=True)
        card = ttk.Frame(outer, style="Surface.TFrame", padding=34)
        card.place(relx=0.5, rely=0.48, anchor="center", width=460)

        ttk.Label(card, text="DATÁCORA", style="Brand.TLabel").pack(anchor="w")
        ttk.Label(card, text="Carga masiva de tareas", style="Title.TLabel").pack(anchor="w", pady=(2, 14))
        ttk.Label(card, text="Acceso exclusivo para Jefes de Mantención y Administradores.", style="Muted.TLabel").pack(anchor="w", pady=(0, 24))

        ttk.Label(card, text="Correo", style="Surface.TLabel").pack(anchor="w")
        ttk.Entry(card, textvariable=self.email, width=42).pack(fill=X, pady=(6, 14))
        ttk.Label(card, text="Contraseña", style="Surface.TLabel").pack(anchor="w")
        password_entry = ttk.Entry(card, textvariable=self.password, width=42, show="*")
        password_entry.pack(fill=X, pady=(6, 20))
        password_entry.bind("<Return>", lambda _event: self.login())
        ttk.Button(card, text="Ingresar", style="Primary.TButton", command=self.login).pack(fill=X)
        ttk.Label(card, textvariable=self.status, style="Muted.TLabel").pack(anchor="w", pady=(18, 0))

    def _build_workspace_view(self):
        header = ttk.Frame(self.workspace_frame, style="Header.TFrame", padding=(22, 16))
        header.pack(fill=X)
        left_header = ttk.Frame(header, style="Header.TFrame")
        left_header.pack(side=LEFT)
        ttk.Label(left_header, text="DATÁCORA", style="HeaderTitle.TLabel").pack(anchor="w")
        ttk.Label(left_header, text="Importador masivo de tareas", style="HeaderMuted.TLabel").pack(anchor="w")
        right_header = ttk.Frame(header, style="Header.TFrame")
        right_header.pack(side=RIGHT)
        ttk.Label(right_header, textvariable=self.session_label, style="HeaderMuted.TLabel").pack(side=LEFT, padx=(0, 12))
        ttk.Button(right_header, text="Cerrar sesión", style="Secondary.TButton", command=self.logout).pack(side=LEFT)

        file_card = ttk.Frame(self.workspace_frame, style="Surface.TFrame", padding=16)
        file_card.pack(fill=X, padx=18, pady=(18, 10))
        ttk.Label(file_card, text="Planilla de tareas", style="Title.TLabel").pack(anchor="w", pady=(0, 10))
        file_bar = ttk.Frame(file_card, style="Surface.TFrame")
        file_bar.pack(fill=X)
        ttk.Entry(file_bar, textvariable=self.excel_path).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(file_bar, text="Seleccionar Excel", style="Secondary.TButton", command=self.choose_file).pack(side=LEFT, padx=8)
        ttk.Button(file_bar, text="Leer y validar", style="Primary.TButton", command=self.load_excel).pack(side=LEFT)
        ttk.Button(file_bar, text="Descargar plantilla", command=self.create_template).pack(side=LEFT, padx=(8, 0))

        actions = ttk.Frame(self.workspace_frame, style="App.TFrame", padding=(18, 0, 18, 8))
        actions.pack(fill=X)
        ttk.Checkbutton(actions, text="Crear solo filas válidas", variable=self.only_valid).pack(side=LEFT)
        ttk.Button(actions, text="Crear tareas", style="Primary.TButton", command=self.create_tasks).pack(side=LEFT, padx=8)
        ttk.Button(actions, text="Exportar resultado", style="Secondary.TButton", command=self.export_results).pack(side=LEFT)
        ttk.Label(actions, textvariable=self.status).pack(side=RIGHT)

        columns = ("fila", "estado", "rbd", "establecimiento", "tecnico", "fecha", "tipo", "prioridad", "mensaje")
        self.tree = ttk.Treeview(self.workspace_frame, columns=columns, show="headings", height=22)
        widths = {"fila": 60, "estado": 110, "rbd": 90, "establecimiento": 230, "tecnico": 230, "fecha": 100, "tipo": 190, "prioridad": 90, "mensaje": 310}
        for column in columns:
            self.tree.heading(column, text=column.capitalize())
            self.tree.column(column, width=widths[column], anchor="w")
        self.tree.pack(fill=BOTH, expand=True, padx=18, pady=(0, 8))

        log_frame = ttk.LabelFrame(self.workspace_frame, text="Detalle")
        log_frame.pack(fill=BOTH, expand=False, padx=18, pady=(0, 14))
        self.log = ttk.Treeview(log_frame, columns=("mensaje",), show="headings", height=5)
        self.log.heading("mensaje", text="Mensaje")
        self.log.column("mensaje", width=1100)
        self.log.pack(fill=BOTH, expand=True)

    def _show_login(self):
        self.workspace_frame.pack_forget()
        self.login_frame.pack(fill=BOTH, expand=True)

    def _show_workspace(self):
        self.login_frame.pack_forget()
        self.workspace_frame.pack(fill=BOTH, expand=True)

    def _login_worker(self):
        self._set_status("Validando credenciales...")
        client = DatacoraClient(self.api_url.get().strip())
        user = client.login(self.email.get().strip(), self.password.get())
        users = client.get_users()
        tasks = client.get_tasks()
        branch_id = user.get("branchId") or user.get("branch_id") or ""
        establishments = client.get_establishments(branch_id=branch_id)
        self.client = client
        self.users = users
        self.tasks = tasks
        self.establishments = establishments
        self.password.set("")
        self.session_label.set(user.get("fullName") or user.get("email") or "Usuario autorizado")
        self._save_config()
        self._refresh_current_metrics()
        self._log(f"Conectado como {self.session_label.get()}. Usuarios: {len(users)}. Tareas visibles: {len(tasks)}")
        self._set_status("Conectado.")
        self._queue_ui(self._show_workspace)

    def logout(self):
        self.client = None
        self.users = []
        self.tasks = []
        self.establishments = []
        self.rows = []
        self._refresh_current_metrics()
        self.password.set("")
        self.session_label.set("")
        self.status.set("Sesión cerrada.")
        if hasattr(self, "tree"):
            for item in self.tree.get_children():
                self.tree.delete(item)
        self._show_login()

    def login(self):
        email = self.email.get().strip()
        password = self.password.get()
        if not email:
            self.status.set("Ingresa el correo.")
            return
        if not password:
            self.status.set("Ingresa la contraseña.")
            return
        self._run_background(self._login_worker)

    def _template_technicians(self):
        technicians = []
        for user in self.users:
            role_name = normalize((user.get("roles") or {}).get("name"))
            status = normalize(user.get("status"))
            if status == "activo" and role_name.startswith("tecnico"):
                name = clean_text(user.get("full_name"))
                if name:
                    technicians.append(name)
        return sorted(set(technicians), key=normalize)

    def create_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="plantilla-carga-masiva-datacora.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tareas"
        sheet.append(MOBILE_TEMPLATE_HEADERS)
        sheet.append(["", "", "", "", "media"])
        sheet.append(["", "", "", "", "media"])

        for index, width in enumerate([20, 34, 14, 28, 14], start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="0B6A43")

        list_sheet = workbook.create_sheet("Listas")
        list_sheet.append(["Técnicos", "Tipos de visita", "Prioridades"])
        technicians = self._template_technicians()
        task_types = MOBILE_TASK_TYPES
        priorities = MOBILE_PRIORITIES
        for index in range(max(len(technicians), len(task_types), len(priorities))):
            list_sheet.append([
                technicians[index] if index < len(technicians) else "",
                task_types[index] if index < len(task_types) else "",
                priorities[index] if index < len(priorities) else ""
            ])
        list_sheet.sheet_state = "hidden"

        technician_end = max(2, len(technicians) + 1)
        type_end = max(2, len(task_types) + 1)
        priority_end = max(2, len(priorities) + 1)
        technician_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"Listas!$A$2:$A${technician_end}",
            allow_blank=True
        )
        type_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"Listas!$B$2:$B${type_end}",
            allow_blank=True
        )
        priority_validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="list",
            formula1=f"Listas!$C$2:$C${priority_end}",
            allow_blank=True
        )
        sheet.add_data_validation(technician_validation)
        sheet.add_data_validation(type_validation)
        sheet.add_data_validation(priority_validation)
        technician_validation.add("B2:B200")
        type_validation.add("D2:D200")
        priority_validation.add("E2:E200")
        workbook.save(path)
        messagebox.showinfo(APP_TITLE, f"Plantilla descargada:\n{path}")

    def _validate_row(self, row):
        rbd = clean_text(row.raw.get("rbd"))
        technician_value = clean_text(row.raw.get("technician"))
        due_date = as_iso_date(row.raw.get("due_date"))
        task_type_value = clean_text(row.raw.get("task_type"))
        task_type = self._mobile_task_type(task_type_value)
        priority = normalize(row.raw.get("priority")) or "media"

        if not due_date:
            row.errors.append("Fecha planificada inválida.")
        if not technician_value:
            row.errors.append("Nombre Técnico requerido.")
        if not rbd:
            row.errors.append("RBD requerido.")
        if not task_type:
            row.errors.append("Tipo Visita inválido.")
        if priority not in MOBILE_PRIORITIES:
            row.errors.append("Prioridad inválida.")

        establishment = None
        if rbd:
            try:
                establishment = self.client.search_establishment(rbd)
            except Exception as exc:
                row.errors.append(f"No se pudo validar RBD: {exc}")
            if not establishment:
                row.errors.append(f"RBD {rbd} no existe.")

        technician = self._find_user(technician_value)
        if not technician:
            row.errors.append(f"Técnico no encontrado: {technician_value}")

        row.payload = {
            "rbd": rbd,
            "assignedTo": technician.get("id") if technician else "",
            "dueDate": due_date,
            "type": task_type or task_type_value,
            "description": f"Tarea cargada masivamente por {self.session_label.get() or self.email.get().strip()}.",
            "priority": priority,
            "status": "urgente" if priority == "alta" else "pendiente",
            "sectionConfig": {}
        }
        row.raw["establishment_name"] = establishment.get("name") if establishment else ""
        row.raw["technician_name"] = technician.get("full_name") if technician else technician_value
        row.status = "Válida" if not row.errors else "Error"

    def _mobile_task_type(self, value):
        normalized_value = normalize(value)
        aliases = {
            "plan de mantencion": "Plan Preventivo Mantención",
            "plan mantencion": "Plan Preventivo Mantención",
            "plan preventivo": "Plan Preventivo Mantención",
            "plan preventivo mantencion": "Plan Preventivo Mantención",
        }
        if normalized_value in aliases:
            return aliases[normalized_value]
        for task_type in MOBILE_TASK_TYPES:
            if normalize(task_type) == normalized_value:
                return task_type
        return ""

    def _setup_styles(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("App.TFrame", background=COLOR_BG)
        style.configure("Surface.TFrame", background=COLOR_SURFACE)
        style.configure("Header.TFrame", background=COLOR_PRIMARY)
        style.configure("Soft.TFrame", background=COLOR_SOFT)
        style.configure("TLabel", background=COLOR_BG, foreground=COLOR_TEXT, font=("Segoe UI", 10))
        style.configure("Surface.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("Segoe UI", 10))
        style.configure("Muted.TLabel", background=COLOR_SURFACE, foreground=COLOR_MUTED, font=("Segoe UI", 9))
        style.configure("Brand.TLabel", background=COLOR_SURFACE, foreground=COLOR_PRIMARY, font=("Segoe UI", 24, "bold"))
        style.configure("Title.TLabel", background=COLOR_SURFACE, foreground=COLOR_TEXT, font=("Segoe UI", 15, "bold"))
        style.configure("PageTitle.TLabel", background=COLOR_BG, foreground=COLOR_TEXT, font=("Segoe UI", 18, "bold"))
        style.configure("PageSub.TLabel", background=COLOR_BG, foreground=COLOR_MUTED, font=("Segoe UI", 10))
        style.configure("StepTitle.TLabel", background=COLOR_SURFACE, foreground=COLOR_PRIMARY, font=("Segoe UI", 10, "bold"))
        style.configure("SoftTitle.TLabel", background=COLOR_SOFT, foreground=COLOR_PRIMARY, font=("Segoe UI", 10, "bold"))
        style.configure("SoftMuted.TLabel", background=COLOR_SOFT, foreground=COLOR_MUTED, font=("Segoe UI", 9))
        style.configure("StatNumber.TLabel", background=COLOR_SOFT, foreground=COLOR_TEXT, font=("Segoe UI", 18, "bold"))
        style.configure("HeaderTitle.TLabel", background=COLOR_PRIMARY, foreground="#ffffff", font=("Segoe UI", 18, "bold"))
        style.configure("HeaderMuted.TLabel", background=COLOR_PRIMARY, foreground="#d8efe5", font=("Segoe UI", 9))
        style.configure("Primary.TButton", background=COLOR_PRIMARY, foreground="#ffffff", font=("Segoe UI", 10, "bold"), padding=(14, 8), borderwidth=0)
        style.map("Primary.TButton", background=[("active", COLOR_PRIMARY_DARK), ("pressed", COLOR_PRIMARY_DARK)])
        style.configure("Secondary.TButton", background=COLOR_SOFT, foreground=COLOR_PRIMARY, font=("Segoe UI", 10, "bold"), padding=(12, 7), borderwidth=0)
        style.map("Secondary.TButton", background=[("active", "#d9eee3"), ("pressed", "#d9eee3")])
        style.configure("TButton", font=("Segoe UI", 10), padding=(10, 6))
        style.configure("TEntry", padding=(8, 7), fieldbackground="#ffffff")
        style.configure("Treeview", font=("Segoe UI", 9), rowheight=30, background="#ffffff", fieldbackground="#ffffff", foreground=COLOR_TEXT)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), foreground=COLOR_TEXT)
        style.configure("TLabelframe", background=COLOR_BG)
        style.configure("TLabelframe.Label", background=COLOR_BG, foreground=COLOR_TEXT, font=("Segoe UI", 10, "bold"))

    def _build_workspace_view(self):
        shell = ttk.Frame(self.workspace_frame, style="App.TFrame")
        shell.pack(fill=BOTH, expand=True)

        sidebar = ttk.Frame(shell, style="Header.TFrame", padding=(18, 22))
        sidebar.pack(side=LEFT, fill="y")
        sidebar.configure(width=230)
        sidebar.pack_propagate(False)
        ttk.Label(sidebar, text="DATÁCORA", style="HeaderTitle.TLabel").pack(anchor="w", pady=(0, 28))
        ttk.Label(sidebar, textvariable=self.session_label, style="HeaderMuted.TLabel").pack(anchor="w")
        ttk.Label(sidebar, text="Jefe de Mantención", style="HeaderMuted.TLabel").pack(anchor="w", pady=(2, 24))
        for item in ("Resumen", "Tareas", "Crear tarea masiva", "Técnicos", "Establecimientos", "Reportes"):
            style = "Primary.TButton" if item == "Crear tarea masiva" else "Secondary.TButton"
            ttk.Button(sidebar, text=item, style=style, command=lambda view=item: self._show_nav_view(view)).pack(fill=X, pady=5)
        ttk.Frame(sidebar, style="Header.TFrame").pack(fill=BOTH, expand=True)
        ttk.Label(sidebar, text="SOSER", style="HeaderTitle.TLabel").pack(anchor="w", pady=(22, 0))

        content = ttk.Frame(shell, style="App.TFrame", padding=(26, 18, 26, 18))
        content.pack(side=LEFT, fill=BOTH, expand=True)
        top = ttk.Frame(content, style="App.TFrame")
        top.pack(fill=X, pady=(0, 18))
        title_box = ttk.Frame(top, style="App.TFrame")
        title_box.pack(side=LEFT)
        ttk.Label(title_box, textvariable=self.page_title, style="PageTitle.TLabel").pack(anchor="w")
        ttk.Label(title_box, textvariable=self.page_subtitle, style="PageSub.TLabel").pack(anchor="w", pady=(4, 0))
        ttk.Button(top, text="Cerrar sesión", style="Secondary.TButton", command=self.logout).pack(side=RIGHT)

        steps = ttk.Frame(content, style="Surface.TFrame", padding=14)
        steps.pack(fill=X, pady=(0, 14))
        for label, detail in (
            ("1  Plantilla", "Descarga el formato"),
            ("2  Archivo", "Selecciona Excel"),
            ("3  Validación", "Revisa errores"),
            ("4  Creación", "Confirma y crea"),
        ):
            step = ttk.Frame(steps, style="Surface.TFrame", padding=(8, 4))
            step.pack(side=LEFT, fill=X, expand=True)
            ttk.Label(step, text=label, style="StepTitle.TLabel").pack(anchor="w")
            ttk.Label(step, text=detail, style="Muted.TLabel").pack(anchor="w")

        body = ttk.Frame(content, style="App.TFrame")
        body.pack(fill=BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=0, minsize=330)
        body.rowconfigure(0, weight=1)
        main_area = ttk.Frame(body, style="App.TFrame")
        main_area.grid(row=0, column=0, sticky="nsew", padx=(0, 16))
        right_area = ttk.Frame(body, style="App.TFrame")
        right_area.grid(row=0, column=1, sticky="ns")
        right_area.configure(width=320)
        right_area.pack_propagate(False)

        self.main_area = main_area
        self._render_import_view()

        preview = ttk.Frame(right_area, style="Surface.TFrame", padding=16)
        preview.pack(fill=X, pady=(0, 14))
        ttk.Label(preview, text="Vista previa de la carga", style="Title.TLabel").pack(anchor="w")
        preview_box = ttk.Frame(preview, style="Soft.TFrame", padding=14)
        preview_box.pack(fill=X, pady=(14, 0))
        ttk.Label(preview_box, textvariable=self.preview_title, style="SoftTitle.TLabel").pack(anchor="w")
        ttk.Label(preview_box, textvariable=self.preview_detail, style="SoftMuted.TLabel", wraplength=260).pack(anchor="w", pady=(8, 0))

        current = ttk.Frame(right_area, style="Surface.TFrame", padding=16)
        current.pack(fill=X, pady=(0, 14))
        ttk.Label(current, text="Estado actual", style="Title.TLabel").pack(anchor="w", pady=(0, 12))
        current_grid = ttk.Frame(current, style="Surface.TFrame")
        current_grid.pack(fill=X)
        self._stat_tile(current_grid, "Pendientes", self.current_pending_count, 0, 0)
        self._stat_tile(current_grid, "Activas", self.current_active_count, 0, 1)
        self._stat_tile(current_grid, "Urgencias", self.current_urgent_count, 1, 0)
        self._stat_tile(current_grid, "Técnicos activos", self.current_technicians_count, 1, 1)
        self._stat_tile(current_grid, "Completadas", self.current_completed_count, 2, 0)
        self._stat_tile(current_grid, "Total tareas", self.current_total_count, 2, 1)

        stats = ttk.Frame(right_area, style="Surface.TFrame", padding=16)
        stats.pack(fill=X)
        ttk.Label(stats, text="Estadísticas de la carga", style="Title.TLabel").pack(anchor="w", pady=(0, 12))
        grid = ttk.Frame(stats, style="Surface.TFrame")
        grid.pack(fill=X)
        self._stat_tile(grid, "Válidas", self.valid_count, 0, 0)
        self._stat_tile(grid, "Con errores", self.error_count, 0, 1)
        self._stat_tile(grid, "Creadas", self.created_count, 1, 0)
        self._stat_tile(grid, "Pendientes", self.pending_count, 1, 1)

        note = ttk.Frame(right_area, style="Soft.TFrame", padding=14)
        note.pack(fill=X, pady=(14, 0))
        ttk.Label(note, text="Creación masiva", style="SoftTitle.TLabel").pack(anchor="w")
        ttk.Label(note, text="Se creará una tarea individual por cada fila válida de la planilla.", style="SoftMuted.TLabel", wraplength=260).pack(anchor="w", pady=(8, 0))

        log_frame = ttk.LabelFrame(right_area, text="Detalle")
        log_frame.pack(fill=BOTH, expand=True, pady=(14, 0))
        self.log = ttk.Treeview(log_frame, columns=("mensaje",), show="headings", height=7)
        self.log.heading("mensaje", text="Mensaje")
        self.log.column("mensaje", width=280)
        self.log.pack(fill=BOTH, expand=True)

    def _clear_main_area(self):
        for child in self.main_area.winfo_children():
            child.destroy()

    def _set_page(self, title, subtitle):
        self.page_title.set(title)
        self.page_subtitle.set(subtitle)

    def _show_nav_view(self, view):
        if view == "Crear tarea masiva":
            self._render_import_view()
        elif view == "Resumen":
            self._render_summary_view()
        elif view == "Tareas":
            self._render_tasks_view()
        elif view == "Técnicos":
            self._render_technicians_view()
        elif view == "Establecimientos":
            self._render_establishments_view()
        elif view == "Reportes":
            self._render_reports_view()

    def _render_import_view(self):
        self._clear_main_area()
        self._set_page("Crear tarea masiva", "Carga una planilla Excel y asigna visitas a técnicos de forma rápida.")

        file_card = ttk.Frame(self.main_area, style="Surface.TFrame", padding=16)
        file_card.pack(fill=X, pady=(0, 14))
        ttk.Label(file_card, text="Planilla de tareas", style="Title.TLabel").pack(anchor="w", pady=(0, 10))
        file_bar = ttk.Frame(file_card, style="Surface.TFrame")
        file_bar.pack(fill=X)
        ttk.Entry(file_bar, textvariable=self.excel_path).pack(side=LEFT, fill=X, expand=True)
        ttk.Button(file_bar, text="Seleccionar Excel", style="Secondary.TButton", command=self.choose_file).pack(side=LEFT, padx=8)
        ttk.Button(file_bar, text="Leer y validar", style="Primary.TButton", command=self.load_excel).pack(side=LEFT)
        ttk.Button(file_bar, text="Descargar plantilla", command=self.create_template).pack(side=LEFT, padx=(8, 0))

        actions = ttk.Frame(self.main_area, style="App.TFrame")
        actions.pack(fill=X, pady=(0, 10))
        ttk.Checkbutton(actions, text="Crear solo filas válidas", variable=self.only_valid).pack(side=LEFT)
        ttk.Button(actions, text="Crear tareas", style="Primary.TButton", command=self.create_tasks).pack(side=RIGHT)
        ttk.Button(actions, text="Exportar resultado", style="Secondary.TButton", command=self.export_results).pack(side=RIGHT, padx=(0, 8))
        ttk.Label(actions, textvariable=self.status).pack(side=LEFT, padx=(16, 0))

        table_card = ttk.Frame(self.main_area, style="Surface.TFrame", padding=10)
        table_card.pack(fill=BOTH, expand=True)
        columns = ("fila", "estado", "rbd", "establecimiento", "tecnico", "fecha", "tipo", "prioridad", "mensaje")
        self.tree = ttk.Treeview(table_card, columns=columns, show="headings", height=18)
        widths = {"fila": 52, "estado": 88, "rbd": 70, "establecimiento": 170, "tecnico": 160, "fecha": 92, "tipo": 130, "prioridad": 74, "mensaje": 170}
        for column in columns:
            self.tree.heading(column, text=column.capitalize())
            self.tree.column(column, width=widths[column], minwidth=45, anchor="w", stretch=True)
        self.tree.pack(fill=BOTH, expand=True)

        help_bar = ttk.Frame(self.main_area, style="Surface.TFrame", padding=(14, 10))
        help_bar.pack(fill=X, pady=(14, 0))
        ttk.Label(help_bar, text="¿Necesitas ayuda?", style="Surface.TLabel").pack(side=LEFT)
        ttk.Label(help_bar, text=" Usa la plantilla oficial para mantener compatibilidad con la app móvil.", style="Muted.TLabel").pack(side=LEFT)
        self._refresh_table()

    def _render_summary_view(self):
        self._clear_main_area()
        self._set_page("Resumen", "Indicadores actuales de tareas, técnicos y carga masiva.")
        card = ttk.Frame(self.main_area, style="Surface.TFrame", padding=18)
        card.pack(fill=X, pady=(0, 14))
        ttk.Label(card, text="Estado actual de Datácora", style="Title.TLabel").pack(anchor="w", pady=(0, 12))
        grid = ttk.Frame(card, style="Surface.TFrame")
        grid.pack(fill=X)
        self._stat_tile(grid, "Pendientes", self.current_pending_count, 0, 0)
        self._stat_tile(grid, "Activas", self.current_active_count, 0, 1)
        self._stat_tile(grid, "Urgencias", self.current_urgent_count, 0, 2)
        self._stat_tile(grid, "Completadas", self.current_completed_count, 1, 0)
        self._stat_tile(grid, "Técnicos activos", self.current_technicians_count, 1, 1)
        self._stat_tile(grid, "Total tareas", self.current_total_count, 1, 2)
        self._render_tasks_table("Últimas tareas visibles", self.tasks[:20])

    def _render_tasks_view(self):
        self._clear_main_area()
        self._set_page("Tareas", "Listado de tareas visibles para el perfil conectado.")
        self._render_tasks_table("Tareas actuales", self.tasks)

    def _render_technicians_view(self):
        self._clear_main_area()
        self._set_page("Técnicos", "Usuarios técnicos registrados en Datácora.")
        technicians = [user for user in self.users if normalize((user.get("roles") or {}).get("name")).startswith("tecnico")]
        rows = []
        for user in technicians:
            rows.append((
                user.get("full_name") or "",
                user.get("email") or "",
                (user.get("roles") or {}).get("name") or "",
                user.get("status") or "",
                (user.get("branches") or {}).get("name") or ""
            ))
        self._render_table("Técnicos", ("Nombre", "Correo", "Perfil", "Estado", "Zona"), rows)

    def _render_establishments_view(self):
        self._clear_main_area()
        self._set_page("Establecimientos", "Muestra de establecimientos disponible desde SQL.")
        rows = []
        for item in self.establishments:
            rows.append((
                item.get("rbd") or "",
                item.get("name") or "",
                item.get("commune") or "",
                item.get("institution_type") or "",
                item.get("branch_name") or ""
            ))
        self._render_table("Establecimientos", ("RBD", "Establecimiento", "Comuna", "Tipo", "Zona"), rows)

    def _render_reports_view(self):
        self._clear_main_area()
        self._set_page("Reportes", "Resumen rápido para seguimiento operativo.")
        by_status = {}
        by_type = {}
        for task in self.tasks:
            by_status[task.get("status") or "Sin estado"] = by_status.get(task.get("status") or "Sin estado", 0) + 1
            by_type[task.get("task_type") or "Sin tipo"] = by_type.get(task.get("task_type") or "Sin tipo", 0) + 1
        rows = [("Estado", key, value) for key, value in sorted(by_status.items())]
        rows.extend(("Tipo", key, value) for key, value in sorted(by_type.items()))
        self._render_table("Reportes", ("Categoría", "Detalle", "Cantidad"), rows)

    def _render_tasks_table(self, title, tasks):
        rows = []
        for task in tasks:
            rows.append((
                task.get("id") or "",
                task.get("status") or "",
                task.get("priority") or "",
                task.get("rbd") or "",
                task.get("establishment_name") or "",
                task.get("assigned_to_full_name") or "",
                task.get("due_date") or "",
                task.get("task_type") or ""
            ))
        self._render_table(title, ("ID", "Estado", "Prioridad", "RBD", "Establecimiento", "Técnico", "Fecha", "Tipo"), rows)

    def _render_table(self, title, columns, rows):
        card = ttk.Frame(self.main_area, style="Surface.TFrame", padding=14)
        card.pack(fill=BOTH, expand=True)
        ttk.Label(card, text=title, style="Title.TLabel").pack(anchor="w", pady=(0, 10))
        table = ttk.Treeview(card, columns=columns, show="headings", height=20)
        for column in columns:
            table.heading(column, text=column)
            table.column(column, width=130, minwidth=70, anchor="w", stretch=True)
        for row in rows:
            table.insert("", END, values=row)
        table.pack(fill=BOTH, expand=True)
        if not rows:
            ttk.Label(card, text="Sin datos para mostrar.", style="Muted.TLabel").pack(anchor="w", pady=(10, 0))

    def _stat_tile(self, parent, label, value_var, row, column):
        tile = ttk.Frame(parent, style="Soft.TFrame", padding=12)
        tile.grid(row=row, column=column, sticky="nsew", padx=5, pady=5)
        parent.columnconfigure(column, weight=1)
        ttk.Label(tile, textvariable=value_var, style="StatNumber.TLabel").pack(anchor="w")
        ttk.Label(tile, text=label, style="SoftMuted.TLabel").pack(anchor="w")

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row in self.rows:
            message = "; ".join(row.errors) if row.errors else row.result or "OK"
            self.tree.insert("", END, values=(
                row.number,
                row.status,
                row.payload.get("rbd", ""),
                row.raw.get("establishment_name", ""),
                row.raw.get("technician_name", ""),
                row.payload.get("dueDate", ""),
                row.payload.get("type", ""),
                row.payload.get("priority", ""),
                message
            ))
        self._refresh_stats()

    def _refresh_stats(self):
        valid = sum(1 for row in self.rows if not row.errors)
        errors = sum(1 for row in self.rows if row.errors)
        created = sum(1 for row in self.rows if row.status == "Creada")
        pending = sum(1 for row in self.rows if row.status in ("Pendiente", "Válida"))
        self.valid_count.set(str(valid))
        self.error_count.set(str(errors))
        self.created_count.set(str(created))
        self.pending_count.set(str(pending))
        if self.rows:
            first = next((row for row in self.rows if not row.errors), self.rows[0])
            self.preview_title.set(first.payload.get("type") or "Carga masiva")
            self.preview_detail.set(
                f"{len(self.rows)} fila(s) leídas. {valid} válida(s), {errors} con error. "
                f"Primer RBD: {first.payload.get('rbd') or '-'}."
            )
        else:
            self.preview_title.set("Sin planilla cargada")
            self.preview_detail.set("Descarga la plantilla, completa las filas y valida antes de crear tareas.")

    def _refresh_current_metrics(self):
        tasks = self.tasks or []
        users = self.users or []
        active_tasks = [task for task in tasks if not self._task_is_completed(task)]
        urgent_tasks = [task for task in active_tasks if self._task_is_urgent(task)]
        pending_tasks = [task for task in active_tasks if not self._task_is_urgent(task)]
        technicians = [user for user in users if self._user_is_active_technician(user)]
        completed = [task for task in tasks if self._task_is_completed(task)]
        self.current_pending_count.set(str(len(pending_tasks)))
        self.current_active_count.set(str(len(active_tasks)))
        self.current_urgent_count.set(str(len(urgent_tasks)))
        self.current_completed_count.set(str(len(completed)))
        self.current_technicians_count.set(str(len(technicians)))
        self.current_total_count.set(str(len(tasks)))

    def _task_is_completed(self, task):
        status = normalize(task.get("status"))
        return status in ("completada", "completed")

    def _task_is_urgent(self, task):
        status = normalize(task.get("status"))
        priority = normalize(task.get("priority"))
        task_type = normalize(task.get("task_type") or task.get("type"))
        return status == "urgente" or priority in ("alta", "high") or "emergencia" in task_type

    def _user_is_active_technician(self, user):
        status = normalize(user.get("status"))
        role = normalize((user.get("roles") or {}).get("name"))
        return status == "activo" and role.startswith("tecnico")


def main():
    root = Tk()
    try:
        ttk.Style().theme_use("vista")
    except Exception:
        pass
    app = TaskImporterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

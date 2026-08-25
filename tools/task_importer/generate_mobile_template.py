from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

headers = ["Fecha planificada", "Nombre TÃ©cnico", "RBD", "Tipo Visita", "Prioridad"]
task_types = ["Plan Preventivo Mantención", "DT", "Mutualidad", "Emergencia", "Acta", "Seremi", "SEC"]
technicians = ["tecnico.rm@soser.cl"]
priorities = ["media", "alta", "baja"]
path = Path(__file__).with_name("Plantilla carga masiva tareas Datacora.xlsx")

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Tareas"
sheet.append(headers)
sheet.append(["", "", "", "", "media"])
sheet.append(["", "", "", "", "media"])

for index, width in enumerate([20, 34, 14, 28, 14], start=1):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

lists = workbook.create_sheet("Listas")
lists.append(["TÃ©cnicos", "Tipos de visita", "Prioridades"])
for index in range(max(len(technicians), len(task_types), len(priorities))):
    lists.append([
        technicians[index] if index < len(technicians) else "",
        task_types[index] if index < len(task_types) else "",
        priorities[index] if index < len(priorities) else ""
    ])
lists.sheet_state = "hidden"

technician_validation = DataValidation(type="list", formula1="Listas!$A$2:$A$2", allow_blank=True)
type_validation = DataValidation(type="list", formula1="Listas!$B$2:$B$8", allow_blank=True)
priority_validation = DataValidation(type="list", formula1="Listas!$C$2:$C$4", allow_blank=True)
sheet.add_data_validation(technician_validation)
sheet.add_data_validation(type_validation)
sheet.add_data_validation(priority_validation)
technician_validation.add("B2:B200")
type_validation.add("D2:D200")
priority_validation.add("E2:E200")

workbook.save(path)
print(path)

